"""script for a simple GUI interface."""

import sys

from pathlib import Path
from typing import Literal

PY_PATHS = " ".join(sys.argv[1:])

for path in PY_PATHS.strip("[]").split(","):
    if path and path not in sys.path:
        sys.path.append(path.strip().replace("\\\\", "\\").replace("'", ""))

#print(f"Python paths: {sys.path}")

import connect_methods
import openpyxl
import openpyxl.styles as excel_styles
import openpyxl.utils as excel_utils
#import PathFunctions
import PySide6.QtCharts as PyCharts
import PySide6.QtCore as PyCore
import PySide6.QtGui as PyGui
import PySide6.QtWidgets as PyWidget
import table_models
import widget_classes

from PathFunctions import PathFunctions
from PySide6.QtCore import QFileSystemWatcher, Qt
from PySide6.QtPrintSupport import QPrintDialog, QPrinter

    #----------------- read file with current lofile location

temp_file = PathFunctions.read_start_file()


class EvalWindow(PyWidget.QMainWindow):


    """ Definition of class EvalWindow which is the main
        GUI build with PySide6 for the dynamic evaluation
        both in table and chart format and options and
        buttons for user interaction
        REMARKS: it will stay as independent APP as soon
        as the PythonPart palette in Allplan is closed
        Closing the app will also stop the event hook and
        the logging process in Allplan
    """


    def __init__(self
                 , window_width:    int
                 , window_height:   int
                 , window_title:    str
                 , row_content:     list
                 , header_naming:   list
                 , param_dict_list: list[dict]):

        """ Class for the GUI of the dynamic evaluation

        Args:
            window_width: width of the GUI window
            window_height: height of the GUI window
            window_title: name of the App window
            row_content: list of lists with table row content
            header_naming: list of the table header names
            param_dict_list: list of dictionaries with attribute/value pairs
            of each relevant Allplan element
        """

        super().__init__()
        self.setWindowTitle(window_title)
        self.resize(window_width, window_height)
        self.row_content = row_content
        self.header_naming = header_naming
        self.param_dict_list = param_dict_list
        self.current_eval_file = str(temp_file)

        #----------------- watch logfile for automated update

        self.eval_file_watcher = QFileSystemWatcher(self)
        self.eval_file_watcher.addPath(self.current_eval_file)
        self.eval_file_watcher.fileChanged.connect(self.eval_file_updated)

        #----------------- menu bar and custom title bar
        #----------------- main menu bar layout and content

        main_menu = self.menuBar()
        main_menu.setObjectName("main_menu")
        main_menu.setFixedHeight(50)
        file_menu = main_menu.addMenu("File")
        export_menu = main_menu.addMenu("Export")
        diagram_menu = main_menu.addMenu("Diagram")

        #----------------- file menu content snapshots

        self.open_snapshot = widget_classes.MenuAction(self, "&Open Snapshot File", "Open saved snapshot file")
        self.save_snapshot = widget_classes.MenuAction(self, "&Save as Snapshot", "Save snapshot of current content")

        file_menu.addAction(self.open_snapshot)
        self.open_snapshot.hovered.connect(lambda: self.show_tooltip_message(self.open_snapshot))
        self.open_snapshot.triggered.connect(lambda: self.handle_file_dialog("open_snapshot"))

        self.open_snapshot.triggered.connect(lambda: self.show_tooltip_message(self.open_snapshot))

        file_menu.addAction(self.save_snapshot)
        self.save_snapshot.hovered.connect(lambda: self.show_tooltip_message(self.save_snapshot))
        self.save_snapshot.triggered.connect(lambda: self.handle_file_dialog("save_snapshot"))
        self.save_snapshot.triggered.connect(lambda: self.show_tooltip_message(self.save_snapshot))

        #----------------- file menu content print

        self.print_content = widget_classes.MenuAction(self, "&Print content", "Print the content")
        file_menu.addAction(self.print_content)
        self.print_content.hovered.connect(lambda: self.show_tooltip_message(self.print_content))
        self.print_content.triggered.connect(lambda: self.show_tooltip_message(self.print_content))
        self.print_content.triggered.connect(self.handle_print_dialog)

        #----------------- export menu content diagram

        self.export_diagram = widget_classes.MenuAction(self, "&Export Diagram"
                                                        , "Export diagram as image")
        export_menu.addAction(self.export_diagram)
        self.export_diagram.hovered.connect(lambda: self.show_tooltip_message(self.export_diagram))

        #----------------- export menu content table

        self.export_diagram.triggered.connect(
            lambda: self.show_tooltip_message(self.export_diagram))
        self.export_diagram.triggered.connect(lambda: self.handle_file_dialog("save_diagram"))
        self.export_table = widget_classes.MenuAction(self, "&Export Table"
                                                        , "Export table as xlsx")
        export_menu.addAction(self.export_table)
        self.export_table.hovered.connect(lambda: self.show_tooltip_message(self.export_table))
        self.export_table.triggered.connect(lambda: self.show_tooltip_message(self.export_table))
        self.export_table.triggered.connect(lambda: self.handle_file_dialog("save_table"))

        #----------------- diagram menu content gradient

        self.isometric_view = widget_classes.MenuAction(self
                                                        , "&Show gradient"
                                                        , "Show diagram with gradient", True)
        self.isometric_view.toggled.connect(self.update_diagram_content)
        self.isometric_view.triggered.connect(
            lambda: self.show_tooltip_message(self.isometric_view))
        self.isometric_view.hovered.connect(lambda: self.show_tooltip_message(self.isometric_view))
        diagram_menu.addAction(self.isometric_view)

        #----------------- diagram menu content legend

        self.draw_legend = widget_classes.MenuAction(self
                                                    , "&Show legend", "Show legend", True)
        self.draw_legend.toggled.connect(self.update_diagram_content)
        self.draw_legend.triggered.connect(lambda: self.show_tooltip_message(self.draw_legend))
        self.draw_legend.hovered.connect(lambda: self.show_tooltip_message(self.draw_legend))
        diagram_menu.addAction(self.draw_legend)

        #----------------- diagram menu content grid

        self.show_grid = widget_classes.MenuAction(self
                                                    , "&Show grid lines"
                                                    , "Show grid lines in diagram", True)
        self.show_grid.toggled.connect(self.update_diagram_content)
        self.show_grid.triggered.connect(lambda: self.show_tooltip_message(self.show_grid))
        self.show_grid.hovered.connect(lambda: self.show_tooltip_message(self.show_grid))
        diagram_menu.addAction(self.show_grid)

        #----------------- diagram menu content labels

        self.show_labels = widget_classes.MenuAction(self
                                                    , "&Show labels", "Show labels", True)
        self.show_labels.toggled.connect(self.update_diagram_content)
        self.show_labels.triggered.connect(lambda: self.show_tooltip_message(self.show_labels))
        self.show_labels.hovered.connect(lambda: self.show_tooltip_message(self.show_labels))
        diagram_menu.addAction(self.show_labels)

        custom_header = widget_classes.TitleBar(40, "Dynamic Evaluation")

        #----------------- bottom status bar

        status_bar = self.statusBar()
        status_bar.setObjectName("status_bar")
        status_bar.showMessage("Dynamic evaluation")


        #----------------- toolbar buttons and components

        self.key_prop = widget_classes.PullDown(200, 30, ["Material", "Trade", "Object_name"])
        key_prop_headline = widget_classes.HeadlineText(200, 30, "Key attribute")

        self.quantity_prop = widget_classes.PullDown(200, 30, ["Volume", "Area", "Length", "Piece"])
        quantity_prop_headline = widget_classes.HeadlineText(200, 30, "Quantity attribute")

        self.diagram_type = widget_classes.PullDown(200, 30, ["Bar","Ring", "Column", "Pie"])
        diagram_type_headline = widget_classes.HeadlineText(200, 30, "Diagram type")

        self.sorting_prop = PyWidget.QWidget()
        self.sorting_prop.setObjectName("sorting_prop")

        self.sorting_option = widget_classes.CheckBox(30, 30)
        self.sorting_prop_pulldown = widget_classes.PullDown(200, 30, ["None", "Material", "Object_name"])

        self.sorting_option.toggled.connect(self.sorting_prop_pulldown.setVisible)
        self.sorting_option.toggled.connect(lambda _: self.update_table_content())
        self.sorting_prop_pulldown.setVisible(self.sorting_option.isChecked())
        self.sorting_prop_pulldown.activated.connect(self.update_table_content)

        self.sorting_prop_layout = PyWidget.QHBoxLayout()
        self.sorting_prop_layout.setContentsMargins(2, 2, 2, 2)
        self.sorting_prop_layout.setSpacing(4)
        self.sorting_prop_layout.addWidget(self.sorting_option)
        self.sorting_prop_layout.addWidget(self.sorting_prop_pulldown)
        self.sorting_prop.setLayout(self.sorting_prop_layout)
        sorting_headline = widget_classes.HeadlineText(200, 30, "Sorting attribute")

        self.key_prop.activated.connect(self.update_table_content)
        self.quantity_prop.activated.connect(self.update_table_content)
        self.diagram_type.activated.connect(self.update_table_content)


        #----------------- bottom bar buttons and components

        update_button = widget_classes.PushButton(120, 30, "Restart dynamic evaluation!")
        close_button = widget_classes.PushButton(80, 30, "Close")
        update_button.clicked.connect(self.update_table_content)
        close_button.clicked.connect(self.close)
        #close_button.clicked.connect(lambda: PathFunctions.delete_folder())


        #----------------- toolbar layout

        canvas_toolbar = PyWidget.QWidget()
        canvas_toolbar.setObjectName("canvas_toolbar")
        canvas_toolbar.setAttribute(Qt.WA_StyledBackground, True)

        toolbar_layout = PyWidget.QGridLayout()
        toolbar_layout.setContentsMargins(8, 4, 8, 4)
        toolbar_layout.setHorizontalSpacing(10)
        toolbar_layout.setVerticalSpacing(6)
        toolbar_layout.addWidget(key_prop_headline, 0, 0)
        toolbar_layout.addWidget(quantity_prop_headline, 0, 1)
        toolbar_layout.addWidget(sorting_headline, 0, 2)
        toolbar_layout.addWidget(diagram_type_headline, 0, 3)
        toolbar_layout.addWidget(self.key_prop, 1, 0)
        toolbar_layout.addWidget(self.quantity_prop, 1, 1)
        toolbar_layout.addWidget(self.sorting_prop, 1, 2)
        toolbar_layout.addWidget(self.diagram_type, 1, 3)

        canvas_toolbar.setLayout(toolbar_layout)

        #----------------- table and diagram content

        self.table_part = PyWidget.QTableView()
        self.table_part.setObjectName("table_part")
        self.table_part.setAttribute(Qt.WA_StyledBackground, True)
        self.table_part.setModel(table_models.MainTableModel(self.row_content, self.header_naming))

        #----------------- table formating and behavior
        self.table_part.setAlternatingRowColors(True)
        self.table_part.setEditTriggers(PyWidget.QAbstractItemView.NoEditTriggers)
        self.table_part.verticalHeader().setDefaultSectionSize(24)
        self.table_part.horizontalHeader().setSectionResizeMode(PyWidget.QHeaderView.Stretch)

        self.chart_part = PyCharts.QChartView()
        self.chart_part.setObjectName("chart_part")
        self.chart_part.setAttribute(Qt.WA_StyledBackground, True)
        self.chart_part.setRenderHint(PyGui.QPainter.Antialiasing)

        part_splitter = PyWidget.QSplitter(Qt.Horizontal)
        part_splitter.addWidget(self.table_part)
        part_splitter.addWidget(self.chart_part)
        part_splitter.setSizes([400, 700])   # initial 40 / 60 split
        part_splitter.setChildrenCollapsible(False)

        #----------------- bottom bar layout

        bottom_bar = PyWidget.QWidget()
        bottom_bar.setObjectName("bottom_bar")
        bottom_bar.setAttribute(Qt.WA_StyledBackground, True)

        bottom_bar_layout = PyWidget.QHBoxLayout()
        bottom_bar_layout.setContentsMargins(8, 4, 8, 4)
        bottom_bar_layout.setSpacing(6)
        bottom_bar_layout.addStretch(1)
        bottom_bar_layout.addWidget(update_button)
        bottom_bar_layout.addWidget(close_button)

        bottom_bar.setLayout(bottom_bar_layout)

        #----------------- main canvas layout and content

        main_canvas = PyWidget.QWidget()
        main_canvas.setObjectName("main_canvas")
        main_canvas.setAttribute(Qt.WA_StyledBackground, True)

        main_canvas_layout = PyWidget.QVBoxLayout()
        main_canvas_layout.setContentsMargins(0, 0, 0, 0)
        main_canvas_layout.setSpacing(0)

        main_canvas_layout.addWidget(canvas_toolbar)
        main_canvas_layout.addWidget(part_splitter, 1)
        main_canvas_layout.addWidget(bottom_bar)

        main_canvas.setLayout(main_canvas_layout)

        #----------------- app window layout and content

        app_window = PyWidget.QWidget()

        window_layout = PyWidget.QVBoxLayout()
        window_layout.setContentsMargins(0, 0, 0, 0)
        window_layout.setSpacing(0)
        window_layout.addWidget(custom_header)
        window_layout.addWidget(main_canvas)
        app_window.setLayout(window_layout)

        self.setCentralWidget(app_window)

        #----------------- fill layout with data

        self.update_diagram_content()

    #----------------- methods for content handling and update
    #----------------- object update method for table

    def update_table_content(self):

        """ method to update the whole app content in creating a
            new table model from the current logfile content both
            when starting the app and the logfile has changed and
            also if the user changes settings in the toolbar
            also triggers the diagramm update method

        """

        new_key_attrib = self.key_prop.currentText().strip()
        new_quantity_attrib = self.quantity_prop.currentText().strip()
        new_diagram_type = self.diagram_type.currentText().strip()
        new_sorting_attrib = self.sorting_prop_pulldown.currentText().strip()

        additional_sorting = (self.sorting_option.isChecked()
                              and self.sorting_prop_pulldown.currentText() != "None")

        if additional_sorting and new_diagram_type in ["Bar", "Column"]:
            self.row_content, self.header_naming = table_models.aggregat_sorted_file(
                self.param_dict_list, new_key_attrib, new_quantity_attrib
                , new_sorting_attrib)
        else:
            self.row_content, self.header_naming = table_models.aggregat_param_file(
                self.param_dict_list, new_key_attrib, new_quantity_attrib)


        self.table_part.setModel(table_models.MainTableModel(
            self.row_content, self.header_naming))
        self.update_diagram_content()

    #----------------- object update method for diagram

    def update_diagram_content(self):

        """ method to update the diagramm content which is based
            on th etable content and only reads the values from there
            will be triggered with the update_table_content method

        """
    #----------------- read table_part data

        key_prop_label = self.header_naming[0]
        quantity_prop_label = self.header_naming[1]

        isometric_view = self.isometric_view.isChecked()
        draw_legend = self.draw_legend.isChecked()
        show_grid = self.show_grid.isChecked()
        show_labels = self.show_labels.isChecked()

        diagram_type = self.diagram_type.currentText().strip()

        additional_sorting = (self.sorting_option.isChecked()
                              and self.sorting_prop_pulldown.currentText() != "None")

    #----------------- if second sorting is set
    #----------------- rearange data to become sets and dicts

        if additional_sorting and diagram_type in ["Bar", "Column"]:
            key_total_value_dict ={}
            sorting_prop_value_dict = {}
            sorting_prop_color_dict = {}

            current_key = ""
            for single_row in self.row_content:
                if str(single_row[0]).strip() != "":
                    current_key = str(single_row[0])
                    key_total_value_dict[current_key] = float(single_row[1])

                sorting_prop_name = str(single_row[2])
                sorting_prop_quantity = float(single_row[3])
                sorting_prop_color = single_row[4]

                if sorting_prop_name not in sorting_prop_value_dict:
                    sorting_prop_value_dict[sorting_prop_name] = {}
                if current_key not in sorting_prop_value_dict[sorting_prop_name]:
                    sorting_prop_value_dict[sorting_prop_name][current_key] = 0.00

                sorting_prop_value_dict[sorting_prop_name][current_key] += sorting_prop_quantity
                sorting_prop_color_dict[sorting_prop_name] = sorting_prop_color

            key_value_list = list(key_total_value_dict.keys())
            quantity_value_list = [key_total_value_dict[key_name]
                                   for key_name in key_value_list]
            color_value_list = [self.row_content[0][4]] * len(key_value_list) if self.row_content else []

    #----------------- if no second sorting is set

        else:
            key_value_list = [str(single_row[0]) for single_row in self.row_content]
            quantity_value_list = [float(single_row[1]) for single_row in self.row_content]
            color_value_list = [single_row[2] for single_row in self.row_content]

        if additional_sorting and diagram_type in ["Bar", "Column"]:
            barset_lable_list = sorted(sorting_prop_value_dict.keys()
                                       , key = lambda name: str(name))
            key_quantity_list = []
            color_value_list = []

            for barset_name in barset_lable_list:
                barset_quantity_list = [
                    sorting_prop_value_dict[barset_name].get(key_name, 0.00)
                                        for key_name in key_value_list]
                key_quantity_list.append(barset_quantity_list)
                color_value_list.append(sorting_prop_color_dict[barset_name])
        else:

            barset_lable_list = key_value_list
            key_quantity_list = []
            for index in range(len(key_value_list)):
                quantity_list = [0.00]*len(key_value_list)
                quantity_list[index] = quantity_value_list[index]
                key_quantity_list.append(quantity_list)

    #----------------- create diagram object
    #----------------- column or bar diagram

        diagram_as_bar = self.diagram_type.currentText() == "Bar"


        diagram_chart_object = PyCharts.QStackedBarSeries()
        if diagram_as_bar:
            diagram_chart_object = PyCharts.QHorizontalStackedBarSeries()

        for barset_lable, value_list, color_value in zip(barset_lable_list
                                                      , key_quantity_list
                                                      , color_value_list):
            chart_set = PyCharts.QBarSet(barset_lable)
            chart_set.append(value_list)


            if isometric_view:
                if diagram_as_bar:
                    gradient = PyGui.QLinearGradient(0, 0, 250, 0)
                else:
                    gradient = PyGui.QLinearGradient(0, 0, 0, 250)
                base_color = PyGui.QColor(color_value[0], color_value[1], color_value[2])
                gradient.setColorAt(0.0, base_color.lighter(150))
                gradient.setColorAt(1.0, base_color.darker(150))
                chart_set.setBrush(PyGui.QLinearGradient(gradient))
            else:
                chart_set.setBrush(PyGui.QColor(color_value[0], color_value[1], color_value[2]))

            chart_set.hovered.connect(lambda hovered
                                      , bar_index
                                      , bar_set=chart_set: self.handle_bar_hovering(hovered, bar_index, bar_set))
            diagram_chart_object.append(chart_set)

        diagram_chart_object.setBarWidth(0.5)

        diagram_chart_part = PyCharts.QChart()
        diagram_chart_part.addSeries(diagram_chart_object)
        diagram_chart_part.setTitle(f"{quantity_prop_label} per {key_prop_label}")
        diagram_chart_part.setAnimationOptions(PyCharts.QChart.SeriesAnimations)

    #----------------- add diagram axis and legend

        diagram_axis_hor = PyCharts.QBarCategoryAxis()
        diagram_axis_hor.append(key_value_list)

        if diagram_as_bar:
            diagram_axis_hor = PyCharts.QValueAxis()
        diagram_chart_part.addAxis(diagram_axis_hor, Qt.AlignBottom)

        diagram_axis_ver = PyCharts.QValueAxis()
        if diagram_as_bar:
            diagram_axis_ver = PyCharts.QBarCategoryAxis()
            diagram_axis_ver.append(key_value_list)
        diagram_chart_part.addAxis(diagram_axis_ver, Qt.AlignLeft)

        diagram_chart_object.attachAxis(diagram_axis_hor)
        diagram_chart_object.attachAxis(diagram_axis_ver)
        if draw_legend:
            diagram_chart_part.legend().setVisible(True)
            diagram_chart_part.legend().setAlignment(Qt.AlignRight)
        else:
            diagram_chart_part.legend().setVisible(False)

        if show_grid:
            diagram_axis_hor.setGridLineVisible(True)
            diagram_axis_ver.setGridLineVisible(True)
        else:
            diagram_axis_hor.setGridLineVisible(False)
            diagram_axis_ver.setGridLineVisible(False)

    #----------------- pie and ring diagram

        pie_chart_part = PyCharts.QPieSeries()
        pie_chart_part.setLabelsVisible(True)

        if self.diagram_type.currentText() == "Ring":
            pie_chart_part.setHoleSize(0.25)

        for key_value, quantity_value, color_value in zip(key_value_list
                                                          , quantity_value_list
                                                          , color_value_list):
            pie_slice = pie_chart_part.append(key_value, quantity_value)
            if isometric_view:
                gradient = PyGui.QRadialGradient(0, 0, 450)
                base_color = PyGui.QColor(color_value[0], color_value[1], color_value[2])
                gradient.setColorAt(0.0, base_color.lighter(150))
                gradient.setColorAt(1.0, base_color.darker(120))
                pie_slice.setBrush(PyGui.QRadialGradient(gradient))
            else:
                pie_slice.setBrush(PyGui.QColor(color_value[0], color_value[1], color_value[2]))

        pie_chart_slices = pie_chart_part.slices()
        for single_slice in pie_chart_slices:
            quantity_kind = self.quantity_prop.currentText()
            if quantity_kind == "Piece":
                value_string = f"{single_slice.value():,.0f}"
            else:
                value_string = f"{single_slice.value():,.2f}"
            single_slice.setLabel(f"{single_slice.label()}\n"
                                  f"{value_string}")
            single_slice.setLabelVisible(show_labels and single_slice.percentage() * 100 >= 3.00)
            single_slice.setLabelPosition(PyCharts.QPieSlice.LabelOutside)
            single_slice.hovered.connect(lambda hovered
                                         , slice = single_slice: slice.setExploded(hovered))
            single_slice.hovered.connect(lambda hovered
                                         , slice = single_slice: self.handle_slice_hovering(hovered, slice))

    #----------------- add diagram labeling and legend

        pie_chart_object = PyCharts.QChart()
        pie_chart_object.addSeries(pie_chart_part)
        pie_chart_object.setTitle(f"{quantity_prop_label} per {key_prop_label}")
        pie_chart_object.setAnimationOptions(PyCharts.QChart.SeriesAnimations)

        if draw_legend:
            pie_chart_object.legend().setVisible(True)
            pie_chart_object.legend().setAlignment(Qt.AlignRight)
        else:
            pie_chart_object.legend().setVisible(False)

    #----------------- add chart to widget

        if self.diagram_type.currentText() in ["Pie", "Ring"]:
            self.chart_part.setChart(pie_chart_object)
        else:

            self.chart_part.setChart(diagram_chart_part)

    #----------------- eval file watching method

    def eval_file_updated(self
                          , eval_file_path: str | Path):

        """ method to inspect the current used logfile for changes
            and updates

        Args:
            eval_file_path: path to the current logfile

        """

        changed_file = str(Path(eval_file_path))
        self.param_dict_list = table_models.read_param_file(changed_file)
        self.update_table_content()
        if changed_file not in self.eval_file_watcher.files():
            self.eval_file_watcher.addPath(changed_file)

    #----------------- satus bar text method

    def show_tooltip_message(self
                             , menu_action: widget_classes.MenuAction):

        """ method to show the tooltip of the menu action in the status bar

        Args:
            menu_action: the menu action for the tooltip
        """

        self.statusBar().showMessage(menu_action.toolTip(), 5000)


    #----------------- window close method

    def closeEvent(self
                   , event):

        """ method to handle the window close event

        Args:
            event: the close event
        """

        PathFunctions.delete_folder()
        PathFunctions.delete_logfile(self.current_eval_file)
        event.accept()


    #----------------- method to handle the file and export dialogs

    def handle_file_dialog(self
                           , action_kind: Literal["open_snapshot"
                                                  , "save_snapshot"
                                                  , "save_table"
                                                  , "save_diagram"] = "open_snapshot"):

        """ combined method for different actions of the menu bar
        Args:
            action_kind: the kind of action to handle

        """
    #----------------- action open snapshot file
        if action_kind == "open_snapshot":
            snapshot_file = connect_methods.menu_action_dialog(action_kind = action_kind)
            if snapshot_file:
                PathFunctions.save_start_file(snapshot_file)
                if self.current_eval_file in self.eval_file_watcher.files():
                    self.eval_file_watcher.removePath(self.current_eval_file)
                self.current_eval_file = snapshot_file
                self.eval_file_watcher.addPath(self.current_eval_file)

                self.param_dict_list = table_models.read_param_file(snapshot_file)
                self.update_table_content()

    #----------------- action save snapshot file
        elif action_kind == "save_snapshot":
            snapshot_file = connect_methods.menu_action_dialog(action_kind = action_kind)
            if snapshot_file:
                with open(self.current_eval_file, "r", encoding="utf-8") as source_file:
                    logfile_line_list = source_file.readlines()

                source_file.close()

                with open(snapshot_file, "w", encoding="utf-8") as snapshot_log_file:
                    snapshot_log_file.writelines(logfile_line_list)

                self.statusBar().showMessage(f"Snapshot saved to {snapshot_file}", 5000)
                snapshot_log_file.close()

    #----------------- action save table as xlsx file
        elif action_kind == "save_table":
            save_excel_file = connect_methods.menu_action_dialog(action_kind=action_kind)
            if save_excel_file:

                header_style = table_models.create_table_style(cell_color = "BFBFBF"
                                                                , border_color = "000000"
                                                                , border_style = "thick"
                                                                , font_size = 14
                                                                , font_weight = True
                                                                , style_name = "header")
                body_style = table_models.create_table_style(cell_color = "FFFFFF"
                                                            , border_color = "000000"
                                                            , border_style = "thin"
                                                            , font_size = 12
                                                            , font_weight = False
                                                            , style_name = "body")


                excel_file = openpyxl.Workbook()
                excel_file.create_sheet("Evaluation table")
                excel_file.add_named_style(header_style)
                excel_file.add_named_style(body_style)
                if "Sheet" in excel_file.sheetnames:
                    del excel_file["Sheet"]

                eval_sheet = excel_file["Evaluation table"]
                column_count = 1
                row_count = 1
                eval_sheet.row_dimensions[row_count].height = 20.00
                for header_name in self.header_naming:
                    column_string = excel_utils.get_column_letter(column_count)
                    eval_sheet.column_dimensions[column_string].width = 30.00
                    header_cell = eval_sheet.cell(row_count
                                                  , column_count, header_name)
                    header_cell.style = header_style
                    column_count += 1

                column_count = 1
                row_count = 2
                eval_sheet.row_dimensions[row_count].height = 15.00
                for row_content in self.row_content:
                    key_cell = eval_sheet.cell(row_count
                                               , column_count, row_content[0])
                    key_cell.style = body_style
                    key_cell.alignment = excel_styles.Alignment(horizontal = "left"
                                                                , vertical = "center")
                    quantity_cell = eval_sheet.cell(row_count
                                                    , column_count + 1, float(f"{row_content[1]:.3f}"))
                    quantity_cell.style = body_style
                    quantity_cell.alignment = excel_styles.Alignment(horizontal = "right"
                                                                    , vertical = "center")
                    color_cell = eval_sheet.cell(row_count
                                                 , column_count + 2, "")
                    color_cell.style = body_style
                    color_string = f"{row_content[2][0]:02X}{row_content[2][1]:02X}{row_content[2][2]:02X}"
                    color_cell.fill = excel_styles.PatternFill(fill_type = "solid"
                                                               , start_color = color_string)
                    column_count = 1
                    row_count += 1

                excel_file.save(save_excel_file)

    #----------------- action save diagram as image
        elif action_kind == "save_diagram":
            save_diagram_file = connect_methods.menu_action_dialog(action_kind = action_kind)
            if save_diagram_file:
                chart_image = self.chart_part.grab()
                chart_image.save(save_diagram_file)

        else:
            self.statusBar().showMessage(f"Unknown action: {action_kind}", 5000)
            return


    #----------------- method to handle the printing dialog

    def handle_print_dialog(self):

        """ general method to handle the printing action
            as it is very complicated the whole process was
            split into different methods for the single steps

        """

        print_engine = QPrinter(QPrinter.HighResolution)
        printing_dialog = QPrintDialog(print_engine, self)
        printing_dialog.setWindowTitle("Print content")
        printing_dialog.setObjectName("printing_dialog")
        if printing_dialog.exec() != PyWidget.QDialog.Accepted:
            self.statusBar().showMessage("Printing canceled", 5000)
            return

        table_print_image = self.paint_table_image()
        chart_print_image = self.chart_part.grab()

        print_painter = PyGui.QPainter(print_engine)

        self.place_print_image(print_engine, print_painter, chart_print_image)
        print_engine.newPage()
        self.place_print_image(print_engine, print_painter, table_print_image)

        print_painter.end()

        self.statusBar().showMessage("Printing...", 5000)

    #----------------- method to resize and place the single content parts

    def place_print_image(self
                        , print_engine: QPrinter
                        , print_painter: PyGui.QPainter
                        , print_image: PyGui.QPixmap):

        """ partly method for the diagramm printing in using
            Pixmap

        Args:
            print_engine: the printer as such
            print_painter: the QT internal painter to use for printing
            print_image: the QT pixmap widget

        """

        printing_area = print_engine.pageRect(QPrinter.DevicePixel)
        content_margin = 20
        content_area = printing_area.adjusted(content_margin, content_margin
                                            , -content_margin, -content_margin)

        # Scale image to fit content area while preserving aspect ratio
        scaled = print_image.scaledToWidth(int(content_area.width() * 0.9)
                                        , Qt.SmoothTransformation)

        # Center horizontally and vertically
        x_position = content_area.left() + (content_area.width() - scaled.width()) // 2
        y_position = content_area.top() + (content_area.height() - scaled.height()) // 2

        print_painter.drawPixmap(x_position, y_position, scaled)


    #----------------- method to draw the table part as image

    def paint_table_image(self) -> PyGui.QPixmap:


        """ partly method for the table printing in using Pixmap
        Returns:
            PyGui.QPixmap: the rendered table as a pixmap

        """

        table_model = self.table_part.model()
        row_count = table_model.rowCount()
        column_count = table_model.columnCount()

        header_height = self.table_part.horizontalHeader().height()
        row_height_list = [self.table_part.rowHeight(row_index) for row_index in range(row_count)]
        column_width_list = [self.table_part.columnWidth(column_index) for column_index in range(column_count)]

        total_table_width = sum(column_width_list)
        total_table_height = header_height + sum(row_height_list)

        # Render at 4x scale for high-quality print output
        dpi_scale = 4
        image_width = total_table_width * dpi_scale
        image_height = total_table_height * dpi_scale

        table_image = PyGui.QImage(image_width, image_height, PyGui.QImage.Format_RGB32)
        table_image.fill(PyGui.QColor(255, 255, 255))

        table_painter = PyGui.QPainter(table_image)
        table_painter.scale(dpi_scale, dpi_scale)
        table_painter.setRenderHint(PyGui.QPainter.Antialiasing)

        x_position = 0
        y_position = 0

        for column_index in range(column_count):
            column_width = column_width_list[column_index]
            header_text = table_model.headerData(column_index, Qt.Horizontal, Qt.DisplayRole)
            table_painter.drawRect(x_position, y_position, column_width, header_height)
            table_painter.drawText(x_position + 6, y_position, column_width - 12, header_height
                                , Qt.AlignVCenter | Qt.AlignLeft, str(header_text))
            x_position += column_width

        y_position = header_height
        for row_index in range(row_count):
            x_position = 0
            row_height = row_height_list[row_index]
            if row_index % 2 == 0:
                row_base_color = PyGui.QColor(255, 255, 255)
            else:
                row_base_color = PyGui.QColor(200, 210, 220)
            for column_index in range(column_count):
                column_width = column_width_list[column_index]
                cell_index = table_model.index(row_index, column_index)
                cell_text = table_model.data(cell_index, Qt.DisplayRole)
                cell_color = table_model.data(cell_index, Qt.BackgroundRole)
                draw_color = cell_color if cell_color is not None else row_base_color

                table_painter.fillRect(x_position, y_position, column_width, row_height, draw_color)
                table_painter.drawRect(x_position, y_position, column_width, row_height)
                table_painter.drawText(x_position + 6, y_position, column_width - 12, row_height
                                    , Qt.AlignVCenter | Qt.AlignLeft, "" if cell_text is None else str(cell_text))
                x_position += column_width
            y_position += row_height

        table_painter.end()
        return PyGui.QPixmap.fromImage(table_image)


    #----------------- method for the mouse hover in the bar charts

    def handle_bar_hovering(self
                            , is_hovered: bool
                            , bar_index: int
                            , hovered_bar: PyCharts.QBarSet) -> None:

        """ method to handle the mouse hovering if the diagram type
            is a bar or column

        Args:
            is_hovered: if the mouse is hovering over the bar
            bar_index: the index of the hovered bar
            hovered_bar: the bar set of the hovered bar

        """
        if not is_hovered or bar_index < 0:
            PyWidget.QToolTip.hideText()
            return

        bar_label = hovered_bar.label()
        bar_value = hovered_bar.at(bar_index)
        value_label = self.style_hover_text(bar_label, bar_value)

        PyWidget.QToolTip.showText(PyGui.QCursor.pos(), value_label, self.chart_part
                                   , msecShowTime = 8000)


    #----------------- method for the mouse hover in the pie charts

    def handle_slice_hovering(self
                            , is_hovered: bool
                            , pie_slice: PyCharts.QPieSlice) -> None:

        """ method to handle the mouse hovering if the diagram type
            is a pie or ring

        Args:
            is_hovered: if the mouse is hovering over the slice
            pie_slice: the pie slice that is being hovered over
        """

        if is_hovered:
            slice_label = pie_slice.label().split("\n")[0]
            slice_value = pie_slice.value()
            value_label = self.style_hover_text(slice_label, slice_value)
            PyWidget.QToolTip.showText(PyGui.QCursor.pos(), value_label, self.chart_part
                                       , msecShowTime = 8000)
        else:
            PyWidget.QToolTip.hideText()

    #----------------- method to style the tooltip text of hoovering

    def style_hover_text(self
                         , lable_text:  str
                         , lable_value: float) -> str:

        """ method to style the tooltip text of hoovering

        Args:
            lable_text: the text of the tooltip
            lable_value: the content of the tooltip

        Returns:
            the styled tooltip text
        """

        quantity_kind = self.quantity_prop.currentText()
        if quantity_kind == "Piece":
            value_string = f"{lable_value:,.0f}"
        else:
            value_string = f"{lable_value:,.2f}"
        return (
            "<div>"
            f"<span style = 'font-family:\"Segoe UI\"; font-size:12px; font-weight:600;'><b>{lable_text}</b></span><br>"
            f"<span style = 'font-family:\"Segoe UI\"; font-size:11px;'>{value_string}</span>"
            "</div>"
        )

    #----------------- FINAL: creation of the main window and start of the app

gui_styles_path = Path(__file__).parent / "gui_styles.qss"
with open(gui_styles_path, "r", encoding="utf-8") as style_file:
    gui_style_sheet = style_file.read()

icon_dir = (Path(__file__).parent / "icons").resolve().as_posix()
gui_style_sheet = gui_style_sheet.replace(
    "url(gui_classes/icons/arrow_down.png)",
    f'url("{icon_dir}/arrow_down.png")'
)
gui_style_sheet = gui_style_sheet.replace(
    "url(gui_classes/icons/checkbox_cross.png)",
    f'url("{icon_dir}/checkbox_cross.png")'
)

eval_app = PyWidget.QApplication(sys.argv)
eval_app.setStyle("Fusion")
eval_app.setStyleSheet(gui_style_sheet)

param_table_list = table_models.read_param_file(temp_file)
param_table, header_list = table_models.aggregat_param_file(param_table_list
                                                            , "Material", "Volume")

eval_app_window = EvalWindow(1200, 800," ", param_table, header_list
                             , param_table_list)

eval_app_window.show()

eval_app.exec()
